# from openpyxl import Workbook
# from openpyxl.styles import Font
#
#
# HEADERS = [
#     "MAIN NAV. ITEM / LAUNCH POINT",
#     "DROPDOWN / NEXT STOP",
#     "FINAL DESTINATION",
#     "PAGE TYPE",
#     "PAGE DESCRIPTION",
#     "KEY SECTIONS / FEATURES",
#     "CONTENT TYPE",
#     "🌈 CONTENT LINK",
#     "🌈 STATUS",
#     "🌈 CLIENT NOTES"
# ]
#
#
# def add_headers(ws):
#     ws.append(HEADERS)
#     for cell in ws[1]:
#         cell.font = Font(bold=True)
#
#
# def normalize_sections(page: dict):
#     return page.get("key features/sections") or page.get("sections") or []
#
#
# def format_sections(sections: list[str] | None):
#     """
#     Render sections as real bullet points (Excel line breaks)
#     """
#     if not sections:
#         return None
#     return "\n".join([f"• {s}" for s in sections])
#
#
# def infer_content_status(description: str | None):
#     if description:
#         return "", ""
#     return "", ""
#
#
# def add_row(
#     ws,
#     main,
#     dropdown,
#     final,
#     page_type,
#     description,
#     sections
# ):
#     content_type, status = infer_content_status(description)
#
#     ws.append([
#         main or None,
#         dropdown or None,
#         final or None,
#         page_type,
#         description,
#         format_sections(sections),
#         content_type,
#         f"{final} Page Copy" if description and final else None,
#         status,
#         None
#     ])
#
#
# def process_page(ws, page):
#     page_name = page["name"]
#
#     # HOME (special case: MAIN + FINAL)
#     if page_name.lower() == "home":
#         add_row(
#             ws=ws,
#             main="Home",
#             dropdown=None,
#             final=None,
#             page_type=page["type"],
#             description=page.get("description"),
#             sections=normalize_sections(page)
#         )
#         return
#
#     # LEVEL 1 — MAIN NAV ITEM ONLY
#     add_row(
#         ws=ws,
#         main=page_name,
#         dropdown=None,
#         final=None,
#         page_type=page["type"],
#         description=page.get("description"),
#         sections=normalize_sections(page)
#     )
#
#     # LEVEL 2 — COLLECTIONS (DROPDOWN ONLY)
#     for sub in page.get("sub_pages") or []:
#         add_row(
#             ws=ws,
#             main=None,
#             dropdown=sub["name"],
#             final=None,
#             page_type=sub["type"],
#             description=sub.get("description"),
#             sections=normalize_sections(sub)
#         )
#
#         # LEVEL 3 — PRODUCTS (FINAL ONLY)
#         for prod in sub.get("sub_sub_pages") or []:
#             add_row(
#                 ws=ws,
#                 main=None,
#                 dropdown=None,
#                 final=prod["name"],
#                 page_type=prod["type"],
#                 description=prod.get("description"),
#                 sections=normalize_sections(prod)
#             )
#
#
# def json_to_excel(sitemap_json: dict, output_path: str):
#     wb = Workbook()
#
#     header_ws = wb.active
#     header_ws.title = "Header"
#
#     footer_ws = wb.create_sheet("Footer")
#
#     add_headers(header_ws)
#     add_headers(footer_ws)
#
#     for page in sitemap_json.get("pages", []):
#         placements = page.get("placement", [])
#
#         if "Header" in placements:
#             process_page(header_ws, page)
#
#         if "Footer" in placements:
#             process_page(footer_ws, page)
#
#     wb.save(output_path)


from openpyxl import Workbook
from openpyxl.styles import Font
from typing import Dict, List, Any, Optional, Tuple

HEADERS = [
    "MAIN NAV. ITEM / LAUNCH POINT",
    "DROPDOWN / NEXT STOP",
    "FINAL DESTINATION",
    "PAGE TYPE",
    "PAGE DESCRIPTION",
    "KEY SECTIONS / FEATURES",
    "CONTENT TYPE",
    "🌈 CONTENT LINK",
    "🌈 STATUS",
    "🌈 CLIENT NOTES"
]


def add_headers(ws):
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def normalize_sections(page: dict) -> List[str]:
    """Retrieves and standardizes the key sections list."""
    # Updated to handle potential variations in key names
    sections = page.get("key_features/sections") or page.get("sections") or page.get("key_sections") or []
    return [s for s in sections if isinstance(s, str)]


def format_sections(sections: List[str] | None) -> Optional[str]:
    """
    Render sections as real bullet points (Excel line breaks)
    """
    if not sections:
        return None
    # Use Excel line break character
    return "\n".join([f"• {s}" for s in sections])


def infer_content_status(description: str | None) -> Tuple[Optional[str], Optional[str]]:
    """
    Placeholder for logic to infer content type and status based on description content.
    For now, return default values based on presence of description.
    """
    if description and len(description.strip()) > 5:
        return "Draft Copy", "In Progress"
    elif description:
        return "Draft Copy", "Needs Review"
    return None, "Awaiting Content"


def add_row(
        ws,
        main: Optional[str],
        dropdown: Optional[str],
        final: Optional[str],
        page_type: Optional[str],
        description: Optional[str],
        sections: List[str]
):
    """Appends a single structured row to the worksheet."""

    # Prioritize the most specific nav item for the content link/status inference
    target_name = final or dropdown or main

    content_type, status = infer_content_status(description)

    ws.append([
        main or None,
        dropdown or None,
        final or None,
        page_type,
        description,
        format_sections(sections),
        content_type,
        f"{target_name} Page Copy" if target_name and content_type else None,
        status,
        None
    ])


# --- CRITICAL CHANGE: Generalized Processing ---

def _process_navigation_section(ws, page: Dict[str, Any], main_parent: Optional[str] = None,
                                dropdown_parent: Optional[str] = None):
    """
    Processes a page and its sub_pages up to three levels deep, assigning
    names to Main, Dropdown, or Final Destination columns based on depth.
    """
    page_name = page["name"]
    is_home = page_name.lower() == "home"

    # Determine the columns based on the current depth
    current_main = None
    current_dropdown = None
    current_final = None

    if is_home:
        current_main = page_name
    elif not main_parent:
        # Level 1: Main Nav Item
        current_main = page_name
    elif not dropdown_parent:
        # Level 2: Dropdown / Next Stop
        current_dropdown = page_name
    else:
        # Level 3 (and deeper): Final Destination
        current_final = page_name

    # Add the row for the current page
    add_row(
        ws=ws,
        main=current_main,
        dropdown=current_dropdown,
        final=current_final,
        page_type=page.get("type"),
        description=page.get("description"),
        sections=normalize_sections(page)
    )

    # Recursively process sub-pages, passing the current page's name as the new parent
    for sub_page in page.get("sub_pages") or []:

        # If we are at Level 1 (current_main), the next level becomes dropdown_parent
        if current_main:
            new_main_parent = current_main
            new_dropdown_parent = None

        # If we are at Level 2 (current_dropdown), the next level becomes final destination.
        # We retain the original L1 parent name.
        elif current_dropdown:
            new_main_parent = main_parent
            new_dropdown_parent = current_dropdown

        # If we are at Level 3 (current_final), we assume it's the FINAL level for this template.
        # We just pass the parents down again.
        else:
            new_main_parent = main_parent
            new_dropdown_parent = dropdown_parent

        # To strictly enforce max 3 levels (Main, Dropdown, Final), we stop recursion if
        # we have already established Main and Dropdown parents.
        if main_parent and dropdown_parent:
            # This means we are processing a L3 sub-page, which is not tracked in this structure
            # For a flat sitemap template, we assume all pages below L3 are also L3 destinations.
            # We reuse the L3 row logic (current_final) above.
            pass

        _process_navigation_section(
            ws=ws,
            page=sub_page,
            main_parent=current_main or main_parent,  # Pass L1 name down
            dropdown_parent=current_dropdown or dropdown_parent  # Pass L2 name down
        )


def json_to_excel(sitemap_json: dict, output_path: str):
    wb = Workbook()

    header_ws = wb.active
    header_ws.title = "Header"

    footer_ws = wb.create_sheet("Footer")

    add_headers(header_ws)
    add_headers(footer_ws)

    # Process all top-level pages
    for page in sitemap_json.get("pages", []):
        placements = page.get("placement", [])

        if "Header" in placements:
            # Start the recursive processing for Header items
            _process_navigation_section(header_ws, page)

        if "Footer" in placements:
            # Start the recursive processing for Footer items
            _process_navigation_section(footer_ws, page)

    wb.save(output_path)

input = {'site': 'Bickford USA', 'pages': [{'name': 'Home', 'type': 'Main Page', 'placement': ['Header'], 'description': None, 'sections': ['Hero Banner', 'Brand Story', 'Testimonials', 'USPs', 'Newsletter', 'Featured Collections'], 'sub_pages': []}, {'name': "Men's", 'type': 'Collection Page', 'placement': ['Header'], 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': [{'name': 'Tops', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}, {'name': 'Bottoms', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}]}, {'name': "Women's", 'type': 'Collection Page', 'placement': ['Header'], 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': [{'name': 'Tops', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}, {'name': 'Bottoms', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}]}, {'name': 'Youth', 'type': 'Collection Page', 'placement': ['Header'], 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': [{'name': 'Tops', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}, {'name': 'Bottoms', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}]}, {'name': 'Accessories', 'type': 'Collection Page', 'placement': ['Header'], 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': [{'name': 'Tops', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}, {'name': 'Bottoms', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}]}, {'name': 'About Us', 'type': 'Main Page', 'placement': ['Footer'], 'description': 'Brand origin, founders, mission, and goals in a concise, general format', 'sections': [], 'sub_pages': []}, {'name': 'Contact Us', 'type': 'Main Page', 'placement': ['Footer'], 'description': 'Contact form, email, phone, and address', 'sections': [], 'sub_pages': []}, {'name': 'FAQ', 'type': 'Main Page', 'placement': ['Footer'], 'description': 'Clearly labeled section for frequently asked questions — specific questions not required', 'sections': ['Frequently Asked Questions'], 'sub_pages': []}, {'name': 'Shop', 'type': 'Utility', 'placement': ['Footer'], 'description': None, 'sections': [], 'sub_pages': [{'name': "Men's", 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}, {'name': "Women's", 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}, {'name': 'Youth', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}, {'name': 'Accessories', 'type': 'Collection Page', 'description': None, 'sections': ['Collection Title', 'Description', 'Product Grid with Size, Color Filters'], 'sub_pages': []}]}, {'name': 'Our Policies', 'type': 'Utility', 'placement': ['Footer'], 'description': None, 'sections': [], 'sub_pages': [{'name': 'Shipping Policy', 'type': 'Policy Page', 'description': 'Clear description of shipping terms and conditions', 'sections': []}, {'name': 'Returns Policy', 'type': 'Policy Page', 'description': 'Clear description of return process and conditions', 'sections': []}, {'name': 'Privacy Policy', 'type': 'Policy Page', 'description': 'Clear description of data collection and privacy practices', 'sections': []}, {'name': 'Terms of Service', 'type': 'Policy Page', 'description': 'Clear description of service terms and user agreements', 'sections': []}]}, {'name': 'Useful Links', 'type': 'Utility', 'placement': ['Footer'], 'description': None, 'sections': [], 'sub_pages': [{'name': 'About', 'type': 'Main Page', 'description': 'Brand origin, founders, mission, and goals in a concise, general format', 'sections': []}, {'name': 'Contact', 'type': 'Main Page', 'description': 'Contact form, email, phone, and address', 'sections': []}, {'name': 'FAQ', 'type': 'Main Page', 'description': 'Clearly labeled section for frequently asked questions — specific questions not required', 'sections': ['Frequently Asked Questions']}]}]}

json_to_excel(input, "example-2.xlsx")